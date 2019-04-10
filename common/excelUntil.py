# Date：9/3/2019
# Name:kobe
# -*- coding:utf-8 -*-

from openpyxl import load_workbook

class ExcelMethod(object):
    def __init__(self, logger, workBookName):
        self.logger = logger
        self.workBookName = workBookName
        try:
            self.wb = load_workbook(workBookName)
        except Exception as e:
            self.logger('获取文件失败: %s' % e)

    def readExcel(self, sheetName, isExcute):
        self.sheetName = sheetName
        self.sheet = self.wb.get_sheet_by_name(sheetName)
        self.max_column = self.sheet.max_column
        dataList = []
        for row in self.sheet.rows:
            tmpList = []
            if row[isExcute].value == 'Y' or row[isExcute].value == 'y':
                for cell in row:
                    tmpList.append(cell.value)
                dataList.append(tmpList)
        return dataList

    def saveExcel(self, row, text):
        try:
            self.sheet.cell(row, self.max_column, text)
            self.wb.save(self.workBookName)
        except:
            self.logger.error('%s 保存失败' % self.workBookName)

if __name__ == '__main__':
    from common.logging_method import LoggingMethod
    import os
    logger = LoggingMethod().getLogger()
    data_path = os.path.dirname(os.path.abspath('.')) + '\\data' + '\\login.xlsx'
    excel = ExcelMethod(logger, data_path)
    user_list = excel.readExcel('Sheet1', 3)

    excel.saveExcel(2, 'pass')
    print(user_list)